# -*- coding: utf-8 -*-
"""
SKLEARN ALGORITHMS

Carlos Rodríguez Outerelo
outerelocarlos@gmail.com
"""

# To install the needed packages
from pip._internal import main as pipmain

'''
pipmain(['install', 'quandl'])
pipmain(['install', 'pandas'])
pipmain(['install', 'numpy'])
pipmain(['install', 'timeit'])
pipmain(['install', 'datetime'])
pipmain(['install', 'sklearn'])
pipmain(['install', 'matplotlib'])
pipmain(['install', 'openpyxl'])
'''

# quandl is used to obtain the needed data
import quandl

# Data is manipulated with Pandas
import pandas as pd

# Math related libraries/functions
import numpy as np
from math import sqrt

# Time related libraries
import timeit
import datetime

# sklearn is used to perform the machine learning tasks at hand
from sklearn.preprocessing import scale
from sklearn.model_selection import train_test_split as tts
from sklearn.linear_model import LinearRegression
from sklearn.neighbors import KNeighborsRegressor
from sklearn.tree import DecisionTreeRegressor
from sklearn.ensemble import AdaBoostRegressor
from sklearn.svm import SVR
from sklearn.metrics import r2_score, mean_squared_error

# matplotlib is used for the plots
import matplotlib.pyplot as plt
from matplotlib import style

# openpyxl is used to save the results in an Excel file
import openpyxl

def main():
    # SVR kernels
    kernels = ['linear', 'poly', 'rbf', 'sigmoid']

    # List of classifiers to use
    clf = [LinearRegression(n_jobs = -1), 
    KNeighborsRegressor(),
    DecisionTreeRegressor(), 
    AdaBoostRegressor(), 
    SVR()]

    # As of now this is the classifier selector, running a loop is too demanding
    clf_choice = 0
    clf = clf[clf_choice]

    # The selected stocks are stored in a list
    stock_list=["WSE/CDPROJEKT",
                "TSE/9684",
    #           "BCHARTS/COINBASEEUR", to check consistency of algorithm bitcoin data
                "BNC3/GWA_LTC",
                "BNC3/GWA_BTC"
                ]

    # The user is prompted for its Quandl token
    token = input("Please enter your Quandl API token: ")

    # This for loop goes through all the stocks in the stock list
    for k in range(len(stock_list)):
        st = timeit.default_timer()
        stock = quandl.get(stock_list[k], authtoken=token) 
        # print(stock) # to check available features
            
        stock = stock[['High', 'Low', 'Open', 'Close']] # Meaningful features    
        
        # Future predictions are NaN -> fillna() avoids NaN-related issues
        stock.fillna(value = -99999, inplace = True)
        runtime = 7 # Number of days to predict (more days = less success)
        
        for kk in stock.columns:
            # A new (shifted) column is created
            # The column is shifted -runtime days, so the last runtime days are NaN
            stock['Shift_%s' % kk] = stock[kk].shift(-runtime)
            
            ft = stock[kk] # The relevant feature is selected
            ft.dropna(inplace = True) # direcly np.array() does not work without dropna(), was hard to debug
            ft = np.array(ft)
            ft = scale(ft) # The feature is scaled (0-1)

            # The tailing days are separated from the feature array
            ft_tail = ft[-runtime:]
            ft = ft[:-runtime]
            
            # The newly created shifted column is selected
            lb = stock['Shift_%s' % kk]
            lb.dropna(inplace = True) # direcly np.array(lb) does not work without dropna(), was hard to debug
            lb = np.array(lb)
            
            print('Stock: ', stock_list[k])
            print('Classifier: ', str(clf))

            # print("Features length: %s; Label length: %s" % (len(ft),len(lb)))
            # If everything has been done properly both lengths should be equal
            
            reduce_var = 100 #another var reductor like test_size, value is example
            prediction = np.empty([reduce_var, runtime]) #faster than np.zeros        
            confidence, confi, rmse = [], [], []
            
            for j in range(reduce_var):
                ft_train, ft_test, lb_train, lb_test = tts(ft.reshape(-1, 1), 
                                                        lb, test_size=0.2)
                
                clf.fit(ft_train, lb_train) #fit = train
                
                confidence.append(clf.score(ft_test, lb_test))
                confi.append(r2_score(lb_test, 
                                    clf.predict(ft_test)))
                rmse.append(sqrt(mean_squared_error(lb_test, 
                                                    clf.predict(ft_test))))
                
                prediction[j] = clf.predict(ft_tail.reshape(-1, 1))
            
            if reduce_var > 1:    
                prediction = prediction.mean(0,dtype=np.float64)
                print('r2 using .score(): ', np.mean(confidence))
                print('r2 using r2_score(): ', np.mean(confi))
                print('RMSE: ', np.mean(rmse))
                    
            stock['Future_%s' % kk] = np.nan
            
            if kk == stock.columns[0]:
                t_stamp = stock.iloc[-1].name #refer 2 panda index & create timestamp
                t_float = t_stamp.timestamp() #from timestamp to float
                t_day = 86400 #daytime in seconds
            
            for i in prediction:
                if kk == stock.columns[0]:
                    t_float += t_day
                    #now we do float 2 timestamp:
                    tomorrow = datetime.datetime.fromtimestamp(t_float)
                    stock.loc[tomorrow] = [np.nan for j in range(len(stock.columns)-1)] + [i]
                    #.loc is panda's index of the column 0: if existing replaces it, if not creates it
                    #[np.nan for j in range(len(stock.columns)-1)] creates NaN for all columns but future
                    # +[i] makes the date of tomorrow associate with the value of tomorrow's prediction which is i
                else:
                    stock.iloc[-runtime+prediction.tolist().index(i),-1] = i
            
            #print(stock)
            
            #Anaconda's Spyder may not work with Matplotlib with its default settings so:
            #Tools > Preferences > IPython console > Graphics
            #change the Graphics backend from Inline to Automatic
            #restart anaconda
            
            # I hid these plots to further revise them

            """
            style.use('ggplot')

            plt.figure(k+1) #if (k > 0 and stock_list[k]) == "BCHARTS/COINBASEEUR" else plt.figure()
            stock[kk].plot()
            stock['Future_%s' % kk].plot()
            plt.title(stock_list[k])
            plt.legend(loc = 2)
            plt.xlabel('Date')
            plt.ylabel('Price')
            plt.show()

            plt.figure((k+1)*100+stock.columns.get_loc(kk)) #the function calls for the column index
            stock[kk].plot()
            stock['Future_%s' % kk].plot()
            plt.title('%s - %s data' % (stock_list[k], kk))
            plt.legend(loc = 2)
            plt.xlabel('Date')
            plt.ylabel('Price')
            plt.show()
            """
        
        # CSV filegen
        file_title = list(stock_list[k])
        for i in range(len(file_title)): 
            if file_title[i] == '/': file_title[i] = '_'
        file_title = "".join(file_title) #http://effbot.org/pyfaq/why-are-python-strings-immutable.htm
        
        stock.to_csv("%s.csv" % file_title, sep = '\t', 
                encoding = 'utf-8', index = True)
        

        # The Excel file is created if k == 0
        if k == 0:
            if clf_choice == 0:
                filepath = "LinearRegression.xlsx"
            elif clf_choice == 1:
                filepath = "KNeighborsRegressor.xlsx"
            elif clf_choice == 2:
                filepath = "DecisionTreeRegressor.xlsx"
            elif clf_choice == 3:
                filepath = "AdaBoostRegressor.xlsx"
            elif clf_choice == 4:
                filepath = "SVR.xlsx"
            openpyxl.Workbook().save(filepath)

        # The data is written to a dedicated sheet in the Excel file
        writer = pd.ExcelWriter(filepath, engine = 'openpyxl', mode = 'a')
        book = openpyxl.load_workbook(filepath)
        writer.book = book
        stock.to_excel(writer, sheet_name = "%s" % file_title, index = True)        

        # Empty sheets are deleted
        if ('Sheet1' in book.sheetnames):
            book.remove(book['Sheet1'])
        elif ('Sheet' in book.sheetnames):
            book.remove(book['Sheet'])

        # The file is saved
        book.save(filepath)
        
        print('Processing time: ', timeit.default_timer() - st, '\n')
    
if __name__ == '__main__':
    main()